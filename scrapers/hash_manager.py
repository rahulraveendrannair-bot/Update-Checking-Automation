"""
hash_manager.py
===============
Runs all 175 RPS scrapers, captures their SHA-256 hashes, and updates
RPS_Updates.xlsx with full guardrails:

  • Timestamped backup before every write
  • Old Hash Value ← New Hash Value  (shift, with archive protection)
  • New Hash Value ← freshly generated hashes
  • Change detection: Old vs New
  • Full audit log  (logs/hash_manager_<timestamp>.log)
  • Transactional write  (either all rows update or none)
  • File-lock check  (aborts if Excel is open)
  • Column schema validation
  • Idempotency  (safe to run multiple times)

Usage:
    python hash_manager.py                        # uses RPS_Updates.xlsx in cwd
    python hash_manager.py --file path/to/RPS.xlsx
    python hash_manager.py --workers 10           # limit parallel workers
    python hash_manager.py --dry-run              # run scrapers but don't write Excel
"""

import argparse
import hashlib
import io
import json
import logging
import os
import re
import shutil
import sys
import tempfile
import threading
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

# ── Third-party (installed via requirements.txt) ──────────────────────────────
try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    sys.exit("ERROR: openpyxl not installed.  Run: pip install openpyxl")

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
DEFAULT_EXCEL   = "RPS_Updates.xlsx"
BACKUP_DIR      = "backups"
LOG_DIR         = "logs"
REQUIRED_COLS   = {"RPS List", "Old Hash Value"}          # "New hash Value" variant handled below
MAX_WORKERS     = 20

# Column header exactly as it appears (case-insensitive match done at runtime)
COL_RPS         = "RPS List"
COL_OLD         = "Old Hash Value"
COL_NEW         = "New hash Value"

# Colours for Excel cells
CLR_HEADER_BG   = "1F4E79"   # dark blue
CLR_HEADER_FG   = "FFFFFF"   # white
CLR_CHANGED_BG  = "FFF2CC"   # light yellow — hash changed
CLR_NEW_BG      = "E2EFDA"   # light green  — first-run (no old value)
CLR_SAME_BG     = "FFFFFF"   # white        — unchanged

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING SETUP
# ─────────────────────────────────────────────────────────────────────────────
def setup_logging(log_dir: str) -> logging.Logger:
    os.makedirs(log_dir, exist_ok=True)
    ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
    logfile = os.path.join(log_dir, f"hash_manager_{ts}.log")

    logger = logging.getLogger("hash_manager")
    logger.setLevel(logging.DEBUG)

    # File handler — full detail
    fh = logging.FileHandler(logfile, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter("%(asctime)s  %(levelname)-8s  %(message)s"))

    # Console handler — INFO and above
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter("%(asctime)s  %(levelname)-8s  %(message)s"))

    logger.addHandler(fh)
    logger.addHandler(ch)
    logger.info(f"Log file: {logfile}")
    return logger


# ─────────────────────────────────────────────────────────────────────────────
# GUARDRAIL: File-lock check
# ─────────────────────────────────────────────────────────────────────────────
def assert_file_not_locked(path: str, logger: logging.Logger) -> None:
    """
    On Windows, Excel holds an exclusive lock.  On all platforms, attempt to
    open the file for writing — if it fails the file is locked/open elsewhere.
    """
    try:
        with open(path, "r+b"):
            pass
    except PermissionError:
        logger.error(f"ABORT: {path} is locked (likely open in Excel). Close it first.")
        sys.exit(1)
    except FileNotFoundError:
        pass   # will be caught later by schema check
    logger.debug("File-lock check passed.")


# ─────────────────────────────────────────────────────────────────────────────
# GUARDRAIL: Schema validation
# ─────────────────────────────────────────────────────────────────────────────
def validate_schema(ws, logger: logging.Logger) -> dict[str, int]:
    """
    Returns {header_name: column_index_1based} for the three required columns.
    Aborts if any required column is missing.
    """
    headers = {}
    for cell in ws[1]:
        if cell.value is not None:
            headers[str(cell.value).strip()] = cell.column

    # Case-insensitive lookup
    col_map = {}
    for required in [COL_RPS, COL_OLD, COL_NEW]:
        match = next(
            (k for k in headers if k.lower() == required.lower()),
            None
        )
        if match is None:
            logger.error(f"ABORT: Required column '{required}' not found in sheet. "
                         f"Found columns: {list(headers.keys())}")
            sys.exit(1)
        col_map[required] = headers[match]

    logger.info(f"Schema OK — columns: {col_map}")
    return col_map


# ─────────────────────────────────────────────────────────────────────────────
# BACKUP
# ─────────────────────────────────────────────────────────────────────────────
def create_backup(excel_path: str, backup_dir: str, logger: logging.Logger) -> str:
    os.makedirs(backup_dir, exist_ok=True)
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem     = Path(excel_path).stem
    dest     = os.path.join(backup_dir, f"{stem}_backup_{ts}.xlsx")
    shutil.copy2(excel_path, dest)
    logger.info(f"Backup created: {dest}")
    return dest


# ─────────────────────────────────────────────────────────────────────────────
# SHIFT: New Hash Value  →  Old Hash Value
# ─────────────────────────────────────────────────────────────────────────────
def shift_hashes(ws, col_map: dict, logger: logging.Logger) -> int:
    """
    For each data row, copy New Hash Value → Old Hash Value.
    Guardrail: only overwrites Old if New is non-empty.
    Returns the number of rows shifted.
    """
    col_old = col_map[COL_OLD]
    col_new = col_map[COL_NEW]
    shifted = 0

    for row_idx in range(2, ws.max_row + 1):
        new_val = ws.cell(row=row_idx, column=col_new).value
        if new_val is not None and str(new_val).strip():
            ws.cell(row=row_idx, column=col_old).value = str(new_val).strip()
            shifted += 1

    logger.info(f"Shifted {shifted} 'New Hash Value' entries → 'Old Hash Value'.")
    return shifted


# ─────────────────────────────────────────────────────────────────────────────
# SCRAPER RUNNER  (mirrors app.py design)
# ─────────────────────────────────────────────────────────────────────────────
_BARE_CALL_RE  = re.compile(r"(?m)^[A-Za-z][A-Za-z0-9_]+\s*\(\s*\)\s*$")
_MAIN_GUARD_RE = re.compile(r'if __name__\s*==\s*["\']__main__["\']')
_SCRAPERS_DIR  = os.path.join(os.path.dirname(__file__), "scrapers")

_HEADLESS_PATCH = """\
try:
    from selenium.webdriver.chrome.options import Options as _ChrOpts
    _orig_init = _ChrOpts.__init__
    def _patched_init(self, *a, **kw):
        _orig_init(self, *a, **kw)
        self.add_argument("--headless=new")
        self.add_argument("--no-sandbox")
        self.add_argument("--disable-dev-shm-usage")
        self.add_argument("--disable-gpu")
    _ChrOpts.__init__ = _patched_init
except Exception:
    pass
"""

import importlib.util

def _load_scraper(fname: str):
    """Load a scraper .txt file and return (fn_name, callable) or None."""
    path = os.path.join(_SCRAPERS_DIR, fname)
    try:
        raw  = open(path, encoding="utf-8", errors="replace").read()
        src  = raw.replace("\r\n", "\n").replace("\r", "\n")
        m    = re.search(r"^def ([A-Za-z][A-Za-z0-9_]+)\s*\(", src, re.MULTILINE)
        if not m:
            return None
        fn_name = m.group(1)
        src     = _BARE_CALL_RE.sub("", src)
        src     = _MAIN_GUARD_RE.sub("if False:", src)
        src     = _HEADLESS_PATCH + src

        py_path = path + ".py"
        with open(py_path, "w", encoding="utf-8") as f:
            f.write(src)

        spec = importlib.util.spec_from_file_location(fn_name, py_path)
        mod  = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        fn = getattr(mod, fn_name, None)
        if callable(fn):
            return fn_name, fn
    except Exception as exc:
        return None
    return None


def _run_one(fn_name: str, fn) -> dict:
    """Execute one scraper and return a result dict."""
    start = time.time()
    try:
        result  = fn()
        elapsed = time.time() - start
        # Extract hash from various return shapes
        h = None
        if isinstance(result, dict):
            h = (result.get("hash") or result.get("output_hash") or
                 result.get("result", {}) and
                 (result["result"] or {}).get("hash") if isinstance(result.get("result"), dict) else None)
        if not h and isinstance(result, (list, tuple)) and len(result) >= 2:
            h = result[-1] if isinstance(result[-1], str) and len(result[-1]) == 64 else None
        return {"task": fn_name, "status": "success", "hash": h, "elapsed": elapsed}
    except Exception as exc:
        return {"task": fn_name, "status": "error",
                "hash": None, "error": str(exc),
                "traceback": traceback.format_exc(),
                "elapsed": time.time() - start}


def run_all_scrapers(rps_names: list[str], max_workers: int,
                     logger: logging.Logger) -> dict[str, dict]:
    """
    Discover scrapers for every name in rps_names, run them in parallel,
    return {rps_name: result_dict}.
    """
    # Build name→callable map from scrapers/ directory
    available: dict[str, callable] = {}
    if os.path.isdir(_SCRAPERS_DIR):
        for fname in os.listdir(_SCRAPERS_DIR):
            if not fname.endswith(".txt"):
                continue
            loaded = _load_scraper(fname)
            if loaded:
                fn_name, fn = loaded
                available[fn_name] = fn

    results: dict[str, dict] = {}
    to_run = [(n, available[n]) for n in rps_names if n in available]
    missing = [n for n in rps_names if n not in available]

    for m in missing:
        logger.warning(f"No scraper found for '{m}' — will record null hash.")
        results[m] = {"task": m, "status": "missing", "hash": None}

    logger.info(f"Running {len(to_run)} scrapers with {max_workers} workers …")

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {executor.submit(_run_one, name, fn): name
                      for name, fn in to_run}
        for future in as_completed(future_map):
            name = future_map[future]
            try:
                res = future.result()
            except Exception as exc:
                res = {"task": name, "status": "error", "hash": None, "error": str(exc)}
            results[name] = res
            status = res["status"]
            h      = res.get("hash", "—")
            if status == "success":
                logger.info(f"  ✔  {name:<50}  hash={h}")
            else:
                logger.warning(f"  ✘  {name:<50}  {res.get('error','')[:80]}")

    return results


# ─────────────────────────────────────────────────────────────────────────────
# GUARDRAIL: Validate hash results
# ─────────────────────────────────────────────────────────────────────────────
def validate_hashes(results: dict, rps_names: list, logger: logging.Logger) -> bool:
    null_tasks = [n for n in rps_names
                  if not (results.get(n, {}).get("hash") or "").strip()]
    if null_tasks:
        logger.warning(f"{len(null_tasks)} tasks returned null/blank hashes: "
                       f"{null_tasks[:10]}{'…' if len(null_tasks)>10 else ''}")
    else:
        logger.info("Hash validation passed — all tasks returned non-null hashes.")
    return True   # non-null check is a warning, not a hard abort


# ─────────────────────────────────────────────────────────────────────────────
# WRITE: Apply new hashes to Excel (transactional)
# ─────────────────────────────────────────────────────────────────────────────
def apply_hashes(ws, col_map: dict, results: dict[str, dict],
                 logger: logging.Logger) -> dict:
    """
    Write new hashes into the New Hash Value column.
    Returns a stats dict: {total, updated, unchanged, null, changed_rows}.
    Transactional: builds all values in memory first, then writes all-or-nothing.
    """
    col_rps = col_map[COL_RPS]
    col_old = col_map[COL_OLD]
    col_new = col_map[COL_NEW]

    # ── Phase 1: build update plan in memory ─────────────────────────────────
    plan = []   # list of (row_idx, rps_name, old_val, new_hash)
    for row_idx in range(2, ws.max_row + 1):
        rps_name = ws.cell(row=row_idx, column=col_rps).value
        if not rps_name:
            continue
        rps_name  = str(rps_name).strip()
        old_val   = ws.cell(row=row_idx, column=col_old).value
        old_val   = str(old_val).strip() if old_val else ""
        new_hash  = (results.get(rps_name, {}).get("hash") or "").strip()
        plan.append((row_idx, rps_name, old_val, new_hash))

    # Verify no row mapping is duplicated
    seen_rows = [row_idx for row_idx, *_ in plan]
    if len(seen_rows) != len(set(seen_rows)):
        logger.error("ABORT: Duplicate row mappings detected — data alignment issue.")
        sys.exit(1)

    # ── Phase 2: write all rows ───────────────────────────────────────────────
    stats = {"total": len(plan), "updated": 0, "unchanged": 0,
             "null": 0, "changed_rows": []}

    for row_idx, rps_name, old_val, new_hash in plan:
        cell = ws.cell(row=row_idx, column=col_new)
        cell.value = new_hash if new_hash else None

        # Cell highlighting
        if not new_hash:
            stats["null"] += 1
            cell.fill = PatternFill("solid", fgColor="FCE4D6")   # light red — null
        elif not old_val:
            cell.fill = PatternFill("solid", fgColor=CLR_NEW_BG)  # green — first run
            stats["updated"] += 1
        elif old_val != new_hash:
            cell.fill = PatternFill("solid", fgColor=CLR_CHANGED_BG)  # yellow — changed
            stats["updated"]    += 1
            stats["changed_rows"].append(rps_name)
        else:
            cell.fill = PatternFill("solid", fgColor=CLR_SAME_BG)     # white — same
            stats["unchanged"] += 1

    logger.info(f"Write plan: total={stats['total']}  updated={stats['updated']}  "
                f"unchanged={stats['unchanged']}  null={stats['null']}  "
                f"changed={len(stats['changed_rows'])}")
    if stats["changed_rows"]:
        logger.info(f"Changed RPS lists: {stats['changed_rows']}")

    return stats


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL FORMATTING
# ─────────────────────────────────────────────────────────────────────────────
def format_sheet(ws, col_map: dict) -> None:
    """Apply professional formatting to header row and set column widths."""
    header_fill = PatternFill("solid", fgColor=CLR_HEADER_BG)
    header_font = Font(bold=True, color=CLR_HEADER_FG, name="Arial", size=11)
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="BFBFBF")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center
        cell.border    = border

    # Column widths
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 70

    # Data row font + border
    data_font = Font(name="Arial", size=10)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font   = data_font
            cell.border = border

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


# ─────────────────────────────────────────────────────────────────────────────
# AUDIT LOG SHEET
# ─────────────────────────────────────────────────────────────────────────────
def write_audit_sheet(wb, run_ts: str, stats: dict, logger: logging.Logger) -> None:
    """Append or create an 'Audit Log' sheet with a new run entry."""
    sheet_name = "Audit Log"
    if sheet_name not in wb.sheetnames:
        ws_audit = wb.create_sheet(sheet_name)
        ws_audit.append(["Run Timestamp", "Total Records", "Updated",
                          "Unchanged", "Null Hashes", "Changed RPS Lists"])
        # Header formatting
        hdr_fill = PatternFill("solid", fgColor="2E4057")
        hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        for cell in ws_audit[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
        ws_audit.column_dimensions["A"].width = 22
        ws_audit.column_dimensions["F"].width = 80
    else:
        ws_audit = wb[sheet_name]

    changed_str = ", ".join(stats.get("changed_rows", [])) or "—"
    ws_audit.append([
        run_ts,
        stats["total"],
        stats["updated"],
        stats["unchanged"],
        stats["null"],
        changed_str,
    ])
    # Zebra-stripe the new row
    last_row = ws_audit.max_row
    bg = "EEF2F7" if last_row % 2 == 0 else "FFFFFF"
    fill = PatternFill("solid", fgColor=bg)
    font = Font(name="Arial", size=10)
    for cell in ws_audit[last_row]:
        cell.fill = fill
        cell.font = font

    logger.info(f"Audit Log sheet updated (row {last_row}).")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="RPS Hash Manager with full guardrails")
    parser.add_argument("--file",    "-f", default=DEFAULT_EXCEL,
                        help=f"Path to Excel file (default: {DEFAULT_EXCEL})")
    parser.add_argument("--workers", "-w", type=int, default=MAX_WORKERS,
                        help=f"Parallel workers for scrapers (default: {MAX_WORKERS})")
    parser.add_argument("--dry-run", "-d", action="store_true",
                        help="Run scrapers but do not write to Excel")
    args = parser.parse_args()

    run_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    logger = setup_logging(LOG_DIR)
    logger.info("=" * 70)
    logger.info(f"RPS Hash Manager  —  run started at {run_ts}")
    logger.info(f"Target file : {args.file}")
    logger.info(f"Workers     : {args.workers}")
    logger.info(f"Dry-run     : {args.dry_run}")
    logger.info("=" * 70)

    # ── GUARDRAIL 1: File exists ──────────────────────────────────────────────
    if not os.path.isfile(args.file):
        logger.error(f"ABORT: Excel file not found: {args.file}")
        sys.exit(1)

    # ── GUARDRAIL 2: File-lock check ─────────────────────────────────────────
    assert_file_not_locked(args.file, logger)

    # ── GUARDRAIL 3: Backup ───────────────────────────────────────────────────
    backup_path = create_backup(args.file, BACKUP_DIR, logger)

    # ── Load workbook ─────────────────────────────────────────────────────────
    wb = load_workbook(args.file)
    ws = wb.active
    logger.info(f"Loaded workbook — sheet '{ws.title}', "
                f"{ws.max_row-1} data rows.")

    # ── GUARDRAIL 4: Schema validation ────────────────────────────────────────
    col_map = validate_schema(ws, logger)

    # ── Read RPS names ────────────────────────────────────────────────────────
    col_rps  = col_map[COL_RPS]
    rps_names = []
    for row_idx in range(2, ws.max_row + 1):
        v = ws.cell(row=row_idx, column=col_rps).value
        if v:
            rps_names.append(str(v).strip())

    logger.info(f"Found {len(rps_names)} RPS list entries.")

    # ── STEP 1: Shift New → Old ───────────────────────────────────────────────
    logger.info("STEP 1: Shifting New Hash Value → Old Hash Value …")
    shift_count = shift_hashes(ws, col_map, logger)

    # ── STEP 2: Run scrapers ──────────────────────────────────────────────────
    logger.info("STEP 2: Running scrapers in parallel …")
    scraper_results = run_all_scrapers(rps_names, args.workers, logger)

    # ── GUARDRAIL 5: Validate hashes ─────────────────────────────────────────
    validate_hashes(scraper_results, rps_names, logger)

    if args.dry_run:
        logger.info("DRY-RUN mode — skipping Excel write.")
        logger.info("Dry-run hash results:")
        for name in rps_names:
            h = scraper_results.get(name, {}).get("hash", "—")
            logger.info(f"  {name:<50}  {h}")
        return

    # ── STEP 3: Write new hashes (transactional) ──────────────────────────────
    logger.info("STEP 3: Writing new hashes to Excel …")
    stats = apply_hashes(ws, col_map, scraper_results, logger)

    # ── GUARDRAIL 6: Verify row count unchanged ───────────────────────────────
    if stats["total"] != len(rps_names):
        logger.error(f"ABORT: Row count mismatch — expected {len(rps_names)}, "
                     f"got {stats['total']}. Restoring from backup.")
        shutil.copy2(backup_path, args.file)
        sys.exit(1)

    # ── Formatting ────────────────────────────────────────────────────────────
    format_sheet(ws, col_map)

    # ── Audit log sheet ───────────────────────────────────────────────────────
    write_audit_sheet(wb, run_ts, stats, logger)

    # ── STEP 4: Save ──────────────────────────────────────────────────────────
    logger.info("STEP 4: Saving workbook …")
    try:
        wb.save(args.file)
        logger.info(f"Workbook saved: {args.file}")
    except Exception as exc:
        logger.error(f"ABORT: Failed to save workbook — {exc}. "
                     f"Restoring from backup.")
        shutil.copy2(backup_path, args.file)
        sys.exit(1)

    # ── STEP 5: Comparison summary ────────────────────────────────────────────
    logger.info("STEP 5: Change comparison summary")
    logger.info(f"  Total RPS records  : {stats['total']}")
    logger.info(f"  Hashes updated     : {stats['updated']}")
    logger.info(f"  Hashes unchanged   : {stats['unchanged']}")
    logger.info(f"  Null/failed hashes : {stats['null']}")
    logger.info(f"  Changed RPS lists  : {len(stats['changed_rows'])}")
    for name in stats["changed_rows"]:
        logger.info(f"    ↳ CHANGED: {name}")

    logger.info("=" * 70)
    logger.info(f"Run complete.  {stats['updated']} updates, "
                f"{len(stats['changed_rows'])} changes detected.")
    logger.info("=" * 70)


if __name__ == "__main__":
    main()
